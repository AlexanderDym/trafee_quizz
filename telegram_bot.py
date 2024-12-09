import time
import logging
import os
from pathlib import Path
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Poll
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, PollAnswerHandler, JobQueue, CallbackContext
from datetime import datetime, time as dt_time
from dotenv import load_dotenv
from openpyxl import load_workbook
from excel_api import initialize_excel, record_user_response, load_authorized_usernames
import csv
from datetime import datetime, timezone
import random


# Load environment variables
load_dotenv(dotenv_path=Path('.') / 'trafee.env')

# Timer for quiz
QUIZ_TIMEOUT_SECONDS = 30

# Global mapping of usernames to chat IDs and their states
user_chat_mapping = {}  # username -> {"chat_id": int, "joined": bool}
poll_participants = {}  # poll_id -> set(user_id)
notified_winners_global = set()  # users who were notified as winners


# Configuration for users and admins
csv_file_path = "registration_log.csv"
authorized_usernames = load_authorized_usernames(csv_file_path)
SUPERADMIN_USERNAME = "Alexander_Dym"
file_path = "updated_bot_list.xlsx"
gifts_file_path = "gifts.xlsx"


# Function to update chat ID mapping
def update_user_chat_mapping(username, chat_id):
    if username and chat_id:
        user_chat_mapping[username] = chat_id

def get_chat_id_by_username(username):
    return user_chat_mapping.get(username)



# Class for quiz questions
class QuizQuestion:
    def __init__(self, question="", answers=None, correct_answer=""):
        self.question = question
        self.answers = answers if answers is not None else []
        self.correct_answer = correct_answer
        self.correct_answer_position = self.__get_correct_answer_position__()

    def __get_correct_answer_position__(self):
        for index, answer in enumerate(self.answers):
            if answer.strip().lower() == self.correct_answer.strip().lower():
                return index
        return -1

# Quiz questions for 7 days
quiz_questions = [
    QuizQuestion("What is an offer? 🎄🎅", ["A call-to-action on a landing page like Hurry up for gifts! 🎁", "A product or service that the advertiser pays for", "Creative content used in advertising, like a holiday card from Santa ❄️"], "A product or service that the advertiser pays for"),
    QuizQuestion("Which button is most commonly used for calls-to-action on Christmas's landing pages? 🎄🎁", ["Read more", "Share", "Buy now"], "Buy now"),
    QuizQuestion("Which social media became the favourite among affiliates during the holiday season thanks to short and dynamic videos? 🎥✨", ["Facebook", "TikTok", "Twitter"], "TikTok"),
    QuizQuestion("Which advertising strategy сan find the most magical ad for the upcoming holidays? 🎅🎄", ["Scaling", "A/B testing", "Retargeting"], "A/B testing"),
    QuizQuestion("What's the metric that measures your earnings per visitor? 🎅💰", ["EPC (Earnings Per Click)", "CTR (Click-Through Rate)", "CPA (Cost Per Action)"], "EPC (Earnings Per Click)"),
    QuizQuestion("What is ROI in affiliate marketing? 🎁📈", ["Ad impressions", "Return on investment and campaign profitability", "Revenue per individual sale"], "Return on investment and campaign profitability"),
    QuizQuestion("What Does CPM Mean in Holiday Advertising? 🎅📊", ["Cost Per Million (cost for one million clicks)", "Cost Per Millisecond (cost for one millisecond)", "Cost Per Mille (cost for one thousand impressions)"], "Cost Per Mille (cost for one thousand impressions)"),
]


def distribute_gifts(day, winners, context):
    """
    Распределяет подарки между победителями и обновляет файлы.

    :param day: Текущий день викторины (1-7).
    :param winners: Список победителей в формате [(user_id, username), ...].
    :param context: CallbackContext для отправки сообщений.
    """
    # Загружаем таблицу подарков
    gifts_wb = load_workbook(gifts_file_path)
    gifts_sheet = gifts_wb[f"Day {day}"]

    # Загружаем файл с пользователями
    users_wb = load_workbook(file_path)
    users_sheet = users_wb[f"Day {day}"]

    # Получаем подарки текущего дня
    available_gifts = []
    for row in gifts_sheet.iter_rows(min_row=2, values_only=True):
        gift_name, quantity, remaining = row
        if remaining > 0:
            available_gifts.append((gift_name, remaining))

    if not available_gifts:
        logging.warning(f"No gifts available for Day {day}.")
        return

    for user_id, username in winners:
        if not available_gifts:
            logging.warning(f"All gifts for Day {day} have been distributed.")
            break

        # Выбираем случайный подарок
        selected_gift = random.choice(available_gifts)
        gift_name, remaining = selected_gift

        # Уменьшаем остаток подарка в таблице
        for row in gifts_sheet.iter_rows(min_row=2):
            if row[0].value == gift_name:
                row[2].value -= 1  # Уменьшаем остаток
                break

        # Записываем победителя в файл пользователей
        winner_found = False
        for row in users_sheet.iter_rows(min_row=2):
            if row[0].value == user_id:  # Сравниваем ID пользователя
                row[4].value = "Winner"  # Отмечаем статус
                row[5].value = gift_name  # Название подарка
                winner_found = True
                break

        if not winner_found:
            # Если записи нет, добавляем новую строку
            users_sheet.append([user_id, username, None, None, "Winner", gift_name])

        # Убираем подарок из доступных, если остаток стал 0
        if remaining - 1 == 0:
            available_gifts = [g for g in available_gifts if g[0] != gift_name]

        # Отправляем пользователю сообщение
        context.bot.send_message(
            chat_id=user_id,
            text=f"🎉 Congratulations, {username}! You won: {gift_name} 🎁"
        )

    # Сохраняем обновления
    gifts_wb.save(gifts_file_path)
    users_wb.save(file_path)

    logging.info(f"Gifts for Day {day} distributed and files updated.")


# Command for superadmin to get the results file
def list_handler(update, context):
    user = update.message.from_user

    if user.username == SUPERADMIN_USERNAME:
        try:
            with open(file_path, 'rb') as file:
                context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename="quiz_results.xlsx")
                update.message.reply_text("👉Here are the current quiz results👈")
        except Exception as e:
            update.message.reply_text(f"Failed to send the file: {str(e)}")
    else:
        update.message.reply_text("⛔ You don't have access to this command.")

# Command to start the quiz for the user
def start_command_handler(update, context):
    user = update.effective_user
    chat_id = update.effective_chat.id
    username = user.username if user.username else "Unknown"

    # Проверка авторизации
    if not is_authorized_user(update):
        logging.warning(f"Unauthorized user @{username} tried to access the bot.")
        context.bot.send_message(chat_id=chat_id, text="⛔ You are not authorized to use this bot.")
        return

    # Check if the user has already started the bot
    if username in user_chat_mapping:
        logging.warning(f"{datetime.now()} - User @{username} tried to press /start again.")
        context.bot.send_message(
            chat_id=chat_id,
            text="You're already in the quiz 👻\n\nThe next question will be tomorrow!\n\nDon't be sneaky 😜."
        )
        return

    # If the user is new, add them to the dictionary
    user_chat_mapping[username] = {"chat_id": chat_id, "joined": False}

    # Send the welcome message
    image_url = "https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz.png"
    welcome_text = (
        "*🎄✨ Welcome to the ultimate holiday quiz challenge! 🎅🎁*\n\n"
        "🔥 From *December 17 to 23*, we'll light up your festive spirit with daily quizzes\n\n"
        "🎯 Answer questions, compete with others, and *grab amazing prizes every day!*\n\n"
        "*🎁 And the grand finale?*\nA special gift waiting for the ultimate champion on Christmas Eve 🎉\n\n"
    )
    context.bot.send_photo(chat_id=chat_id, photo=image_url, caption=welcome_text, parse_mode="Markdown")

    keyboard = [[InlineKeyboardButton("🚀 Join the Quiz", callback_data="participate")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.bot.send_message(chat_id=chat_id, text="Click 'Join the Quiz' to get started.\n\nLet the fun begin! 🎉", reply_markup=reply_markup)




def handle_poll_timeout(context):
    poll_id = context.job.context['poll_id']
    day = context.job.context['day']

    # Список пользователей, которые уже ответили
    answered_users = poll_participants.get(poll_id, set())

    # Загружаем Excel и проверяем, кто уже записан
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    # Создаём множество с ID пользователей, записанных в Excel
    recorded_users = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if isinstance(row[0], int)}

    for username, user_data in user_chat_mapping.items():
        chat_id = user_data.get("chat_id")  # Извлекаем chat_id
        user_id = chat_id  # В данном случае chat_id — это user_id

        if user_id in answered_users or user_id in recorded_users:
            # Пользователь уже ответил, пропускаем
            logging.info(f"User {username} has already answered the question. Timeout skipped.")
            continue

        # Если пользователь не ответил, уведомляем и записываем результат
        try:
            context.bot.send_message(chat_id=chat_id, text="⏰ Time's up!\n\nYour response was not counted 🥵.")
            response_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            record_user_response(file_path, user_id=user_id, username=username, day=day, response_time=response_time, result=False)
        except Exception as e:
            logging.error(f"Failed to notify user {username} (Chat ID: {chat_id}): {e}")

    # Переходим к выбору победителей
    select_winners(context, day)



def notify_users_about_next_day(context):
    day = context.job.context.get('day', 0)  # Получаем текущий день
    next_day = day + 1  # Завтрашний день

    # Убедимся, что номер дня не превышает 7
    if next_day > 7:
        next_day -= 7

    for username, user_data in user_chat_mapping.items():
        if user_data.get("joined"):
            chat_id = user_data.get("chat_id")
            try:
                context.bot.send_message(
                    chat_id=chat_id,
                    text=f"🎄 Reminder! Tomorrow is Day {next_day} of our 7-day holiday giveaway! 🎁✨\n\n"
                         "Don’t miss your chance to win more amazing prizes.\n\n"
                         "🕒 The fun starts at 15:00 UTC sharp, and we’ll send you a reminder 3 minutes before "
                         "to make sure you're ready to shine! 🌟 See you there!"
                )
                logging.info(f"Reminder for next day sent to {username} (Chat ID: {chat_id})")
            except Exception as e:
                logging.error(f"Failed to send next day reminder to {username}: {e}")


    

def select_winners(context, day):
    global notified_winners_global
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    correct_users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == "Верно" and row[4] != "Winner":
            correct_users.append((row[0], row[1]))  # (user_id, username)

    if not correct_users:
        logging.info(f"No correct answers for Day {day + 1}. No winners selected.")
    else:
        # Выбираем 5 победителей
        winners = random.sample(correct_users, min(5, len(correct_users)))

        # Распределяем подарки
        distribute_gifts(day + 1, winners, context)  # Передаем context в distribute_gifts

        # Отправляем уведомления победителям
        for winner in winners:
            user_id, username = winner
            #if user_id not in notified_winners_global:
                #try:
                    #context.bot.send_message(
                        #chat_id=user_id,
                        #text=f"🎉 Congratulations!\n\nYou are the winner of the day! 🏆✨"
                    #)
                    #logging.info(f"Winner notification sent to user ID: {user_id}")
                    #notified_winners_global.add(user_id)
                #except Exception as e:
                    #logging.error(f"Failed to send winner notification to user ID: {user_id}: {e}")

    wb.save(file_path)
    logging.info(f"Winners for Day {day + 1} have been recorded in the Excel sheet.")

    # Планируем отправку напоминания через 5 секунд
    context.job_queue.run_once(
        notify_users_about_next_day,
        when=5,
        context={'day': day + 1}
    )
    logging.info(f"Reminder for next day scheduled in 5 seconds.")



# Callback for participating in quiz
def participate_handler(update, context):
    query = update.callback_query
    query.answer()

    user = query.from_user
    username = user.username if user.username else "Unknown"

    # Проверка авторизации
    if not is_authorized_user(update):
        logging.warning(f"Unauthorized user @{username} tried to join the quiz.")
        query.edit_message_text(text="⛔ You are not authorized to join this quiz.")
        return

    # Проверка: если пользователь уже нажимал Join Quiz
    if user_chat_mapping.get(username, {}).get("joined"):
        logging.info(f"User @{username} tried to join the quiz again.")
        query.edit_message_text(text="You are already in the quiz! 🚀")
        return

    # Отметить пользователя как присоединившегося
    user_chat_mapping[username] = {
        "chat_id": query.message.chat_id,
        "joined": True
    }
    logging.info(f"User @{username} joined the quiz for the first time.")
    query.edit_message_text(text="Welcome to the quiz! 🎉")




# Function to send quiz question
def send_daily_quiz(context, day):
    logging.info(f"Preparing to send quiz for day {day + 1}")

    if day < len(quiz_questions):
        question = quiz_questions[day]

        if not user_chat_mapping:
            logging.warning("⚠️ No users registered for the quiz. Skipping.")
            return

        for username, user_data in user_chat_mapping.items():
            chat_id = user_data.get("chat_id")
            try:
                # Логируем отправку сообщения
                logging.info(f"Sending quiz question to @{username} (Chat ID: {chat_id})")
                context.bot.send_message(
                    chat_id=chat_id,
                    text="⚡ Today's quiz question:"
                )
                add_quiz_question(context, question, chat_id, day)
            except Exception as e:
                logging.error(f"Failed to send quiz question to @{username} (Chat ID: {chat_id}): {e}")

        # Обновляем текущий день
        next_day = (day + 1) % len(quiz_questions)
        context.dispatcher.bot_data['current_day'] = next_day
    else:
        logging.error(f"Day {day + 1} is out of range for questions.")




# Function to notify users about the quiz
def notify_users_about_quiz(context):
    for username, user_data in user_chat_mapping.items():  # Используем user_chat_mapping вместо joined_users
        chat_id = user_data["chat_id"]
        try:
            context.bot.send_message(
                chat_id=chat_id,
                text="The quiz will start in 5 minutes!🔔\n\n"
                     "🔥Get ready!"
            )
            logging.info(f"Reminder sent to {username} (Chat ID: {chat_id})")
        except Exception as e:
            logging.error(f"Failed to send reminder to {username}: {e}")




# Function to send quiz question to user
def add_quiz_question(context, quiz_question, chat_id, day):
    message = context.bot.send_poll(
        chat_id=chat_id,
        question=quiz_question.question,
        options=quiz_question.answers,
        type=Poll.QUIZ,
        correct_option_id=quiz_question.correct_answer_position,
        open_period=QUIZ_TIMEOUT_SECONDS,
        is_anonymous=False,
        explanation="Don't be sad"
    )
    
    # Сохраняем данные опроса
    context.bot_data.update({message.poll.id: {'chat_id': message.chat.id, 'day': day}})
    
    # Планируем таймаут
    context.job_queue.run_once(
        handle_poll_timeout,
        when=QUIZ_TIMEOUT_SECONDS,
        context={'poll_id': message.poll.id, 'day': day}
    )


# Poll answer handler
def poll_handler(update, context):
    poll_answer = update.poll_answer
    user_id = poll_answer.user.id  # Get user_id
    poll_id = poll_answer.poll_id  # Get poll_id
    selected_option_id = poll_answer.option_ids[0]  # Get the user's selected answer

    # Retrieve poll data
    poll_data = context.bot_data.get(poll_id, {})
    day = poll_data.get('day', 0)
    question = quiz_questions[day]
    correct_option_id = question.correct_answer_position
    is_correct = (selected_option_id == correct_option_id)

    # Log poll answer
    logging.info(f"Poll answer received. User: {user_id}, Poll ID: {poll_id}, Selected Option: {selected_option_id}, Correct: {is_correct}")

    # Initialize poll participants if not already done
    if poll_id not in poll_participants:
        poll_participants[poll_id] = set()

    # Add the user to poll_participants
    poll_participants[poll_id].add(user_id)
    logging.info(f"User {user_id} added to poll_participants for poll_id {poll_id}. Current participants: {poll_participants[poll_id]}")

    # Record the result in the table
    response_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    username = poll_answer.user.username if poll_answer.user.username else "Unknown"
    record_user_response(file_path, user_id=user_id, username=username, day=day, response_time=response_time, result=is_correct)

    # Notify the user
    if is_correct:
        context.bot.send_message(
            chat_id=user_id,
            text="🎉 Congratulations, your answer is correct!\n\n🏁 We will now wait for all participants to complete the game.\n\n✨ After that, we will randomly select 20 winners from those who answered correctly.\n\n☘️ Good luck!"
        )
    else:
        context.bot.send_message(
            chat_id=user_id,
            text="❌ Oops, that’s the wrong answer!\n\nBut don’t give up!\n\n🤗 Try again tomorrow."
        )


def is_authorized_user(update):
    user = update.effective_user
    username = user.username

    logging.info(f"Checking authorization for @{username}")

    if username == SUPERADMIN_USERNAME:
        logging.info(f"User @{username} is the superadmin. Access granted.")
        return True

    try:
        # Читаем CSV
        with open(csv_file_path, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)

            logging.info(f"Reading {csv_file_path} for @{username}.")

            for row in reader:
                logging.debug(f"Checking row: {row}")
                # Проверяем соответствие Telegram Username
                if row.get("Telegram Username") == username:
                    logging.info(f"Access granted for user @{username}")
                    return True

        logging.info(f"User @{username} not found in the list.")
    except FileNotFoundError:
        logging.warning(f"⚠️ File {csv_file_path} not found. No user authorization possible.")
    except Exception as e:
        logging.error(f"⚠️ Error reading file {csv_file_path}: {e}")

    # Если пользователь не найден
    logging.warning(f"Access denied for user @{username}")
    return False



# Main function
def main():
    global file_path
    initialize_excel(file_path=file_path)

    TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
    if not TELEGRAM_TOKEN:
        logging.error("TELEGRAM_TOKEN is not set. Exiting.")
        return

    updater = Updater(TELEGRAM_TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start_command_handler))
    dp.add_handler(CommandHandler("list", list_handler))
    dp.add_handler(CallbackQueryHandler(participate_handler, pattern="participate"))
    dp.add_handler(PollAnswerHandler(poll_handler))

    # Инициализация текущего дня
    dp.bot_data['current_day'] = 0  # Начинаем с 0-го дня

    # Логирование времени сервера
    logging.info(f"Current server UTC time: {datetime.now(timezone.utc)}")

    # Планирование задач
    job_queue = updater.job_queue
    # Уведомление за 5 минут до викторины
    job_queue.run_daily(
        notify_users_about_quiz,
        time=dt_time(18, 29),  # Уведомление в 14:55 по UTC
    )
    logging.info("JobQueue task for quiz notifications added at 14:55 UTC.")

    # Планирование самой викторины
    job_queue.run_daily(
        lambda context: send_daily_quiz(context, dp.bot_data['current_day']),
        time=dt_time(18, 3)  # Викторина в 15:00 по UTC
    )
    logging.info("JobQueue task for quiz scheduling added at 15:00 UTC.")
    updater.start_polling()
    logging.info("Bot started in polling mode")


logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
main()