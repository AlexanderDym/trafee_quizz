import logging
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    Filters,
    ConversationHandler,
    CallbackContext,
    CallbackQueryHandler,
)

import sys
from pathlib import Path
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))
from db_api.connection import Database

from dotenv import load_dotenv
load_dotenv(dotenv_path=Path('.') / '.env')
TELEGRAM_TOKEN = os.getenv("REGISTRATOR_TOKEN")

if not TELEGRAM_TOKEN:
    raise ValueError("Токен не найден. Убедитесь, что переменная REGISTRATOR_TOKEN задана в файле trafee.env.")

database = Database()

main_bot_link = "https://t.me/TrafeeQuizBot"  # Ссылка на основного бота

# # Выделение пользователя зелёным цветом
# def mark_user_as_registered(username):
#     """Отмечает пользователя зелёным цветом в таблице user_list.xlsx"""
#     if not os.path.exists(allowed_users_file):
#         logging.error(f"Файл {allowed_users_file} отсутствует. Невозможно отметить пользователя.")
#         return
    
#     try:
#         wb = load_workbook(allowed_users_file)
#         ws = wb.active

#         # Поиск строки с указанным username
#         username_cleaned = username.strip().lower()
#         for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
#             cell_value = str(row[0].value).strip().lower()
#             if cell_value == username_cleaned:
#                 green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
#                 for cell in row:
#                     cell.fill = green_fill
#                 wb.save(allowed_users_file)  # Сохраняем изменения
#                 logging.info(f"Пользователь {username} успешно отмечен зелёным.")
#                 return

#         logging.warning(f"Пользователь {username} не найден в таблице.")
#     except Exception as e:
#         logging.error(f"Ошибка при сохранении регистрации пользователя {username}: {e}")

# # Логирование регистрации в CSV
# def log_registration(username, telegram_username):
#     """Записывает данные регистрации в CSV-файл"""
#     data = {
#         "Trafee Username": username,
#         "Telegram Username": telegram_username,
#         "Registration Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#     }
#     file_exists = os.path.exists(registration_log_file)

#     try:
#         with open(registration_log_file, mode="a", encoding="utf-8") as file:
#             if not file_exists:
#                 file.write("Trafee Username,Telegram Username,Registration Date\n")  # Заголовок
#             file.write(f"{data['Trafee Username']},{data['Telegram Username']},{data['Registration Date']}\n")
#         logging.info(f"Регистрация пользователя {username} записана в CSV.")
#     except Exception as e:
#         logging.error(f"Ошибка при записи данных в CSV: {e}")

# Функция для старта регистрации
def start(update: Update, context: CallbackContext):
    # URL изображения
    image_url = "https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/sign_up.png"

    # Отправляем изображение с текстом
    context.bot.send_photo(
        chat_id=update.effective_chat.id,
        photo=image_url,
        caption=(
            "👋🎅🎄Welcome aboard!\n\n"
            "Please enter your *Trafee username*✨\n\n"
            "So we can verify your access and get you started 🚀"
        ),
        parse_mode="Markdown"
    )
    return 1



# Проверка username
def check_username(update: Update, context: CallbackContext):
    username = update.message.text.strip()

    already_registered = database.get_participant_by_telegram_id(str(update.message.from_user.id))

    if already_registered:
        # Отправляем сообщение с картинкой для уже зарегистрированных пользователей
        context.bot.send_photo(
            chat_id=update.effective_chat.id,
            photo="https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/already_done.png",
            caption=(
                f"✅ {already_registered.trafee_username}, you're already registered!\n\n"
                f"Here's the link to [the main bot]({main_bot_link}) to continue your journey. 🚀"
            ),
            parse_mode="Markdown"
        )
        return ConversationHandler.END

    participant = database.register_user_by_trafee_username(username)

    if participant:
        if participant.telegram_id:
            # Отправляем сообщение с картинкой для пользователей, которые уже зарегистрированы
            context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo="https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/already_done.png",
                caption=(
                    f"✅ {username}, you're already registered!\n\n"
                    f"Here's the link to [the main bot]({main_bot_link}) to continue your journey. 🚀"
                ),
                parse_mode="Markdown"
            )
            return ConversationHandler.END

        else:
            # Отправляем сообщение с картинкой для успешной регистрации
            context.bot.send_photo(
                chat_id=update.effective_chat.id,
                photo="https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/reg_done.png",
                caption=(
                    f"🎉 Congratulations!\n\nYou’ve successfully registered.\n"
                    f"Jump into [the main bot]({main_bot_link}) to start exploring! 🌟"
                ),
                parse_mode="Markdown"
            )
            participant.telegram_id = update.message.from_user.id
            participant.telegram_username = update.message.from_user.name
            participant.name = update.message.from_user.first_name
            database.update_participant(participant)
            return ConversationHandler.END

    else:
        # Если пользователь не найден, отправляем сообщение с картинкой
        keyboard = [[InlineKeyboardButton("🔄 Try Again", callback_data="retry")]]
        reply_markup = InlineKeyboardMarkup(keyboard)

        context.bot.send_photo(
            chat_id=update.effective_chat.id,
            photo="https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/wrong-1.png",
            caption=(
                "😔 Sorry, but your Trafee username isn’t in the list.\n\n"
                "If you think this is a mistake, please contact your manager for assistance.☘️"
            ),
            reply_markup=reply_markup
        )
        return 1




# Обработка кнопки "Попробовать ещё раз"
def retry_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    
    # Отправка сообщения с картинкой
    context.bot.send_photo(
        chat_id=query.message.chat.id,
        photo="https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/Trafee_quiz/try_again.png",  # Укажите URL вашей картинки
        caption=(
            "🚀 No worries!\n\n"
            "Please enter your *Trafee username* again, and let's try one more time."
        ),
        parse_mode="Markdown"
    )
    return 1


# Cancel registration
def cancel(update: Update, context: CallbackContext):
    update.message.reply_text(
        "🚫 Registration has been canceled.\n\nIf you want to try again, just type /start.\n\n We'll be here waiting for you! 😊",
        parse_mode="Markdown"
    )
    return ConversationHandler.END


# Основная функция
def main():
    try:
        updater = Updater(TELEGRAM_TOKEN, use_context=True)
        dp = updater.dispatcher

        # ConversationHandler для регистрации
        conv_handler = ConversationHandler(
            entry_points=[CommandHandler("start", start)],
            states={
                1: [
                    MessageHandler(Filters.text & ~Filters.command, check_username),
                    CallbackQueryHandler(retry_handler, pattern="^retry$")
                ],
            },
            fallbacks=[CommandHandler("cancel", cancel)],
        )

        dp.add_handler(conv_handler)

        # # Команда для отправки таблицы
        # dp.add_handler(CommandHandler("user_list", send_user_list))

        # Запуск бота
        updater.start_polling()
        logging.info("Bot started successfully!")

    except Exception as e:
        logging.error(f"Error starting bot: {str(e)}")


# logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
# main()
