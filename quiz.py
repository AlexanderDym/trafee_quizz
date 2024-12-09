import logging
from datetime import datetime

# Telegram API
from telegram import Poll
from telegram.ext import CallbackContext

# OpenPyXL for working with Excel files
from openpyxl import load_workbook

# Custom modules
from excel_api import record_user_response
from winners import select_winners
from config import file_path, QUIZ_TIMEOUT_SECONDS  # Конфигурационные параметры

# Global variables (предполагается, что они определены глобально или передаются через контекст)
from shared import user_chat_mapping, poll_participants





def send_daily_quiz(context, day):
    logging.info(f"Preparing to send quiz for day {day + 1}")

    if day < len(quiz_questions):
        question = quiz_questions[day]

        if not user_chat_mapping:
            logging.warning("⚠️ No users registered for the quiz. Skipping.")
            return

        for username, user_data in user_chat_mapping.items():
            chat_id = user_data.get("chat_id")
            try:
                # Логируем отправку сообщения
                logging.info(f"Sending quiz question to @{username} (Chat ID: {chat_id})")
                context.bot.send_message(
                    chat_id=chat_id,
                    text="⚡ Today's quiz question:"
                )
                add_quiz_question(context, question, chat_id, day)
            except Exception as e:
                logging.error(f"Failed to send quiz question to @{username} (Chat ID: {chat_id}): {e}")

        # Обновляем текущий день
        next_day = (day + 1) % len(quiz_questions)
        context.dispatcher.bot_data['current_day'] = next_day
    else:
        logging.error(f"Day {day + 1} is out of range for questions.")

def add_quiz_question(context, quiz_question, chat_id, day):
    message = context.bot.send_poll(
        chat_id=chat_id,
        question=quiz_question.question,
        options=quiz_question.answers,
        type=Poll.QUIZ,
        correct_option_id=quiz_question.correct_answer_position,
        open_period=QUIZ_TIMEOUT_SECONDS,
        is_anonymous=False,
        explanation="Don't be sad"
    )
    
    # Сохраняем данные опроса
    context.bot_data.update({message.poll.id: {'chat_id': message.chat.id, 'day': day}})
    
    # Планируем таймаут
    context.job_queue.run_once(
        handle_poll_timeout,
        when=QUIZ_TIMEOUT_SECONDS,
        context={'poll_id': message.poll.id, 'day': day}
    )

    
def handle_poll_timeout(context):
    poll_id = context.job.context['poll_id']
    day = context.job.context['day']

    # Список пользователей, которые уже ответили
    answered_users = poll_participants.get(poll_id, set())

    # Загружаем Excel и проверяем, кто уже записан
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    # Создаём множество с ID пользователей, записанных в Excel
    recorded_users = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if isinstance(row[0], int)}

    for username, user_data in user_chat_mapping.items():
        chat_id = user_data.get("chat_id")  # Извлекаем chat_id
        user_id = chat_id  # В данном случае chat_id — это user_id

        if user_id in answered_users or user_id in recorded_users:
            # Пользователь уже ответил, пропускаем
            logging.info(f"User {username} has already answered the question. Timeout skipped.")
            continue

        # Если пользователь не ответил, уведомляем и записываем результат
        try:
            context.bot.send_message(chat_id=chat_id, text="⏰ Time's up!\n\nYour response was not counted 🥵.")
            response_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            record_user_response(file_path, user_id=user_id, username=username, day=day, response_time=response_time, result=False)
        except Exception as e:
            logging.error(f"Failed to notify user {username} (Chat ID: {chat_id}): {e}")

    # Переходим к выбору победителей
    select_winners(context, day)


# Class for quiz questions
class QuizQuestion:
    def __init__(self, question="", answers=None, correct_answer=""):
        self.question = question
        self.answers = answers if answers is not None else []
        self.correct_answer = correct_answer
        self.correct_answer_position = self.__get_correct_answer_position__()

    def __get_correct_answer_position__(self):
        for index, answer in enumerate(self.answers):
            if answer.strip().lower() == self.correct_answer.strip().lower():
                return index
        return -1

# Quiz questions for 7 days
quiz_questions = [
    QuizQuestion("What is an offer? 🎄🎅", ["A call-to-action on a landing page like Hurry up for gifts! 🎁", "A product or service that the advertiser pays for", "Creative content used in advertising, like a holiday card from Santa ❄️"], "A product or service that the advertiser pays for"),
    QuizQuestion("Which button is most commonly used for calls-to-action on Christmas's landing pages? 🎄🎁", ["Read more", "Share", "Buy now"], "Buy now"),
    QuizQuestion("Which social media became the favourite among affiliates during the holiday season thanks to short and dynamic videos? 🎥✨", ["Facebook", "TikTok", "Twitter"], "TikTok"),
    QuizQuestion("Which advertising strategy сan find the most magical ad for the upcoming holidays? 🎅🎄", ["Scaling", "A/B testing", "Retargeting"], "A/B testing"),
    QuizQuestion("What's the metric that measures your earnings per visitor? 🎅💰", ["EPC (Earnings Per Click)", "CTR (Click-Through Rate)", "CPA (Cost Per Action)"], "EPC (Earnings Per Click)"),
    QuizQuestion("What is ROI in affiliate marketing? 🎁📈", ["Ad impressions", "Return on investment and campaign profitability", "Revenue per individual sale"], "Return on investment and campaign profitability"),
    QuizQuestion("What Does CPM Mean in Holiday Advertising? 🎅📊", ["Cost Per Million (cost for one million clicks)", "Cost Per Millisecond (cost for one millisecond)", "Cost Per Mille (cost for one thousand impressions)"], "Cost Per Mille (cost for one thousand impressions)"),
]