import time
import logging
import os
from pathlib import Path
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Poll
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, PollAnswerHandler, JobQueue
from datetime import datetime, time as dt_time
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
import csv
from datetime import datetime, timezone
import random


# Load environment variables
load_dotenv(dotenv_path=Path('.') / 'trafee.env')

# Timer for quiz
QUIZ_TIMEOUT_SECONDS = 15

# Global mapping of usernames to chat IDs
user_chat_mapping = {}
poll_participants = {}  # poll_id -> set(user_id)
user_participation = {} # Обработка нажатия команды старт
quiz_participation= {} # Обработка нажатия Участия в викторине
notified_winners_global = set()

# Список призов для каждого дня викторины
prizes = [
    "🎁 Сегодня разыгрывается подписка на Spotify Premium на 1 месяц!",
    "🎁 Сегодня разыгрывается подарочная карта Amazon на $20!",
    "🎁 Сегодня разыгрывается подписка на Netflix на 1 месяц!",
    "🎁 Сегодня разыгрывается эксклюзивный мерч от нашей компании!",
    "🎁 Сегодня разыгрывается подписка на YouTube Premium на 1 месяц!",
    "🎁 Сегодня разыгрывается книга-бестселлер в электронном формате!",
    "🎁 Сегодня разыгрывается сертификат на доставку еды на $15!"
]

# Function to update chat ID mapping
def update_user_chat_mapping(username, chat_id):
    if username and chat_id:
        user_chat_mapping[username] = chat_id

def get_chat_id_by_username(username):
    return user_chat_mapping.get(username)

# Function to load authorized usernames from CSV
def load_authorized_usernames(file_path):
    usernames = []
    try:
        with open(file_path, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                if "Telegram Username" in row:
                    usernames.append(row["Telegram Username"])
    except FileNotFoundError:
        logging.warning(f"⚠️ File {file_path} not found. Authorized user list will be empty.")
    except Exception as e:
        logging.error(f"⚠️ Error reading file {file_path}: {e}")
    return usernames

# Configuration for users and admins
csv_file_path = "registration_log.csv"
authorized_usernames = load_authorized_usernames(csv_file_path)
SUPERADMIN_USERNAME = "Alexander_Dym"
file_path = "updated_bot_list.xlsx"

# Initialize the Excel file
def initialize_excel():
    if not os.path.exists(file_path):
        wb = Workbook()
        for i in range(1, 8):
            sheet = wb.create_sheet(title=f"Day {i}")
            headers = ["User ID", "Username", "Response Time", "Correct Answer"]
            sheet.append(headers)
        wb.remove(wb["Sheet"])
        wb.save(file_path)
        logging.info(f"Excel file initialized with sheets for each quiz day at {file_path}")

# Class for quiz questions
class QuizQuestion:
    def __init__(self, question="", answers=None, correct_answer=""):
        self.question = question
        self.answers = answers if answers is not None else []
        self.correct_answer = correct_answer
        self.correct_answer_position = self.__get_correct_answer_position__()

    def __get_correct_answer_position__(self):
        for index, answer in enumerate(self.answers):
            if answer.strip().lower() == self.correct_answer.strip().lower():
                return index
        return -1

# Quiz questions for 7 days
quiz_questions = [
    QuizQuestion("Which of these is a popular affiliate marketing model?", ["Cost per click (CPC)", "Cost per lead (CPL)", "Pay per hire (PPH)"], "Cost per lead (CPL)"),
    QuizQuestion("Какая столица Франции?", ["Лондон", "Берлин", "Париж"], "Париж"),
    QuizQuestion("Сколько планет в Солнечной системе?", ["7", "8", "9"], "8"),
    QuizQuestion("Какое самое большое животное на земле?", ["Слон", "Синий кит", "Жираф"], "Синий кит"),
    QuizQuestion("Сколько секунд в одной минуте?", ["60", "100", "120"], "60"),
    QuizQuestion("Какой элемент обозначается символом O?", ["Кислород", "Водород", "Азот"], "Кислород"),
    QuizQuestion("Кто является создателем теории относительности?", ["Ньютон", "Эйнштейн", "Галилей"], "Эйнштейн"),
]

# Record user response in Excel
def record_user_response(user_id, username, day, response_time, result):
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    sheet = wb[sheet_name]

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    result_text = "Верно" if result else "Неверно"
    result_fill = green_fill if result else red_fill

    # Проверяем, существует ли уже запись для этого пользователя
    user_found = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == user_id:  # Проверяем по user_id
            row[2].value = response_time  # Обновляем время ответа
            row[3].value = result_text  # Обновляем результат
            row[3].fill = result_fill  # Применяем цвет
            user_found = True
            break

    if not user_found:
        # Если записи нет, добавляем новую
        new_row = [user_id, username, response_time, result_text]
        sheet.append(new_row)

        # Применяем цвет заливки к новой строке
        for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=4):
            if cell[3].value == "Верно":
                cell[3].fill = green_fill
            elif cell[3].value == "Неверно":
                cell[3].fill = red_fill

    wb.save(file_path)
    logging.info(f"Результат для пользователя {username} записан: {result_text}")




# Command for superadmin to get the results file
def list_handler(update, context):
    user = update.message.from_user

    if user.username == SUPERADMIN_USERNAME:
        try:
            with open(file_path, 'rb') as file:
                context.bot.send_document(chat_id=update.effective_chat.id, document=file, filename="quiz_results.xlsx")
                update.message.reply_text("Here are the current quiz results.")
        except Exception as e:
            update.message.reply_text(f"Failed to send the file: {str(e)}")
    else:
        update.message.reply_text("⛔ You don't have access to this command.")

# Command to start the quiz for the user
def start_command_handler(update, context):
    user = update.effective_user
    chat_id = update.effective_chat.id
    username = user.username if user.username else "Unknown"

    # Проверяем, запускал ли пользователь уже бота
    if username in user_participation:
        # Логируем повторный запуск
        logging.warning(f"{datetime.now()} - Пользователь @{username} пытался повторно нажать /start.")
        # Отправляем сообщение пользователю
        context.bot.send_message(
            chat_id=chat_id,
            text="Вы уже участвуете в викторине. Следующий вопрос будет завтра! Не шалите 😜."
        )
        return

    # Если пользователь новый, добавляем его в словарь
    user_participation[username] = {"participated": True, "timestamp": datetime.now()}
    
    # Отправляем стандартное приветственное сообщение
    update_user_chat_mapping(username, chat_id)
    image_url = "https://mailer.ucliq.com/wizz/frontend/assets/files/customer/kd629xy3hj208/img/girl-wearing-santa-s-hat-%281%29.jpg"
    welcome_text = (
        "🎄🎅 Добро пожаловать на наш праздничный розыгрыш от 17 по 24 декабря 🎅🎄\n\n"
        "✨ Мы будем задавать вам вопросы каждый день и разыгрывать призы среди участников\n"
        "🎁 Главный приз — уникальный подарок в канун Рождества\n\n"
        "Нажмите кнопку ниже, чтобы участвовать в викторине и выиграть 🎉"
    )
    context.bot.send_photo(chat_id=chat_id, photo=image_url, caption=welcome_text)

    keyboard = [[InlineKeyboardButton("📝 Участвовать в викторине", callback_data="participate")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    context.bot.send_message(chat_id=chat_id, text="Нажмите 'Участвовать в викторине', чтобы начать.", reply_markup=reply_markup)


def handle_poll_timeout(context):
    poll_id = context.job.context['poll_id']
    day = context.job.context['day']

    # Список пользователей, которые уже ответили
    answered_users = poll_participants.get(poll_id, set())

    # Загружаем Excel и проверяем, кто уже записан
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    # Список пользователей, уже записанных в Excel
    recorded_users = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}

    for username, chat_id in user_chat_mapping.items():
        user_id = chat_id  # Если chat_id соответствует user_id
        if user_id in answered_users or user_id in recorded_users:
            # Пользователь уже ответил, пропускаем
            logging.info(f"Пользователь {username} уже ответил на вопрос. Таймаут пропущен.")
            continue

        # Если пользователь не ответил, уведомляем и записываем результат
        context.bot.send_message(chat_id=chat_id, text="⏰ Увы, время вышло! Ваш ответ не был засчитан.")
        response_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        record_user_response(user_id=user_id, username=username, day=day, response_time=response_time, result=False)

    # Переходим к выбору победителей
    select_winners(context, day)




def select_winners(context, day):
    global notified_winners_global
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    # Собираем правильные ответы
    correct_users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == "Верно":  # Колонка "Correct Answer"
            correct_users.append(row)

    # Выбираем победителей
    if len(correct_users) > 5:  # Изменено на 5
        winners = random.sample(correct_users, 5)  # Изменено на 5
    else:
        winners = correct_users

    # Текущий приз
    prize_message = prizes[day] if day < len(prizes) else "🎁 Приз за сегодня будет объявлен позже!"

    # Отправляем сообщения победителям
    for winner in winners:
        user_id = winner[0]  # Предполагается, что ID пользователя в первом столбце
        if user_id not in notified_winners_global:
            context.bot.send_message(
                chat_id=user_id,
                text=f"🎉 Поздравляем! Вы в числе победителей дня! 🏆✨ Свяжитесь с вашим менеджером для получения приза.\n\n{prize_message}"
            )
            notified_winners_global.add(user_id)  # Добавляем в глобальный список

    # Остальным правильным участникам отправляем утешительное сообщение
    for user in correct_users:
        if user not in winners and user[0] not in notified_winners_global:
            user_id = user[0]
            context.bot.send_message(
                chat_id=user_id,
                text="Ваш ответ верный, но в этот раз удача была не на вашей стороне. Попробуйте завтра! 🎯"
            )
            notified_winners_global.add(user_id)


# Callback for participating in quiz
def participate_handler(update, context):
    query = update.callback_query
    query.answer()
    
    user = query.from_user
    chat_id = query.message.chat_id
    username = user.username if user.username else "Unknown"

    # Проверяем, участвовал ли пользователь уже
    if username in quiz_participation:
        # Если пользователь уже участвовал, отправляем сообщение
        context.bot.send_message(
            chat_id=chat_id,
            text="Вы уже участвуете в сегодняшней викторине. Ждите новый вопрос завтра! 😊"
        )
        logging.warning(f"{datetime.now()} - Пользователь @{username} пытался повторно нажать 'Участвовать в викторине'.")
        return

    # Если пользователь новый, регистрируем его участие
    quiz_participation[username] = {"participated": True, "timestamp": datetime.now()}
    
    # Отправляем сообщение о готовности к викторине
    context.bot.send_message(
        chat_id=chat_id,
        text="🕒 Спасибо за участие! Ожидайте, вопрос появится через минуту. Не пропустите уведомление! 🎉"
    )


# Function to send quiz question
# Function to send quiz question
def send_daily_quiz(context, day):
    logging.info(f"Preparing to send quiz for day {day + 1}")

    if day < len(quiz_questions):
        question = quiz_questions[day]

        if not user_chat_mapping:
            logging.warning("⚠️ No users registered for the quiz. Skipping.")
            return

        for username, chat_id in user_chat_mapping.items():
            add_quiz_question(context, question, chat_id, day)

        # Update current day
        next_day = (day + 1) % len(quiz_questions)
        context.dispatcher.bot_data['current_day'] = next_day

        # Log when the next question will be sent
        next_quiz_time = context.job_queue.jobs()[1].next_t.replace(tzinfo=None)
        logging.info(f"Следующий вопрос (день {next_day + 1}) будет отправлен {next_quiz_time}.")
    else:
        logging.error(f"Day {day + 1} is out of range for questions.")



# Function to notify users about the quiz
def notify_users_about_quiz(context):
    """Уведомляет всех участников о том, что викторина начнется через минуту."""
    if not user_chat_mapping:
        logging.warning("⚠️ No users to notify about the quiz.")
        return

    for username, chat_id in user_chat_mapping.items():
        try:
            context.bot.send_message(
                chat_id=chat_id,
                text="🔔 Викторина начнется через 1 минуту! Готовьтесь!"
            )
            logging.info(f"Sent notification to {username} (Chat ID: {chat_id})")
        except Exception as e:
            logging.error(f"Failed to notify user {username} (Chat ID: {chat_id}): {e}")




# Function to send quiz question to user
def add_quiz_question(context, quiz_question, chat_id, day):
    message = context.bot.send_poll(
        chat_id=chat_id,
        question=quiz_question.question,
        options=quiz_question.answers,
        type=Poll.QUIZ,
        correct_option_id=quiz_question.correct_answer_position,
        open_period=QUIZ_TIMEOUT_SECONDS,
        is_anonymous=False,
        explanation="Ты не глупый. Просто так бывает"
    )
    
    # Сохраняем данные опроса
    context.bot_data.update({message.poll.id: {'chat_id': message.chat.id, 'day': day}})
    
    # Планируем таймаут
    context.job_queue.run_once(
        handle_poll_timeout,
        when=QUIZ_TIMEOUT_SECONDS,
        context={'poll_id': message.poll.id, 'day': day}
    )


# Poll answer handler
def poll_handler(update, context):
    poll_answer = update.poll_answer
    user_id = poll_answer.user.id
    poll_id = poll_answer.poll_id
    selected_option_id = poll_answer.option_ids[0]

    poll_data = context.bot_data.get(poll_id, {})
    day = poll_data.get('day', 0)
    question = quiz_questions[day]
    correct_option_id = question.correct_answer_position
    is_correct = (selected_option_id == correct_option_id)

    response_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    username = poll_answer.user.username if poll_answer.user.username else "Unknown"

    # Добавляем пользователя в poll_participants, если его еще нет
    if poll_id not in poll_participants:
        poll_participants[poll_id] = set()
    poll_participants[poll_id].add(user_id)

    # Записываем результат в таблицу
    record_user_response(user_id=user_id, username=username, day=day, response_time=response_time, result=is_correct)

    # Отправляем сообщение пользователю
    if is_correct:
        context.bot.send_message(
            chat_id=user_id,
            text="Поздравляем, ваш ответ правильный! 🎉 Теперь мы подождем, пока все участники завершат игру. После этого мы случайным образом выберем 20 победителей среди тех, кто ответил верно. Удачи!"
        )
    else:
        context.bot.send_message(
            chat_id=user_id,
            text="❌ Упс, это неправильный ответ! Но не сдавайтесь! 🎯 Попробуйте завтра снова."
        )

# Check if user is authorized
def is_authorized_user(update):
    user = update.effective_user
    return user.username == SUPERADMIN_USERNAME or user.username in authorized_usernames

# Main function
def main():
    initialize_excel()

    TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
    if not TELEGRAM_TOKEN:
        logging.error("TELEGRAM_TOKEN is not set. Exiting.")
        return

    updater = Updater(TELEGRAM_TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start_command_handler))
    dp.add_handler(CommandHandler("list", list_handler))
    dp.add_handler(CallbackQueryHandler(participate_handler, pattern="participate"))
    dp.add_handler(PollAnswerHandler(poll_handler))

    # Инициализация текущего дня
    dp.bot_data['current_day'] = 0  # Начинаем с 0-го дня

    # Логирование времени сервера
    logging.info(f"Current server UTC time: {datetime.now(timezone.utc)}")

    # Планирование задач
    job_queue = updater.job_queue
    # Уведомление за 5 минут до викторины
    job_queue.run_daily(
        notify_users_about_quiz,
        time=dt_time(15, 18),  # Уведомление в 14:55 по UTC
    )
    logging.info("JobQueue task for quiz notifications added at 14:55 UTC.")

    # Планирование самой викторины
    job_queue.run_daily(
        lambda context: send_daily_quiz(context, dp.bot_data['current_day']),
        time=dt_time(15, 19)  # Викторина в 15:00 по UTC
    )
    logging.info("JobQueue task for quiz scheduling added at 15:00 UTC.")
    updater.start_polling()
    logging.info("Bot started in polling mode")

if __name__ == "__main__":
    logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
    main()