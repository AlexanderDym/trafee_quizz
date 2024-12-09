import logging
import random
from openpyxl import load_workbook
from telegram.ext import CallbackContext

# Убедитесь, что пути к файлам заданы корректно
from pathlib import Path
import os
from dotenv import load_dotenv
from bots.config import file_path, gifts_file_path
from notifications import notify_users_about_next_day

# Load environment variables (если используется)
load_dotenv(dotenv_path=Path('.') / 'trafee.env')



def select_winners(context, day):
    global notified_winners_global
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"
    sheet = wb[sheet_name]

    correct_users = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == "Верно" and row[4] != "Winner":
            correct_users.append((row[0], row[1]))  # (user_id, username)

    if not correct_users:
        logging.info(f"No correct answers for Day {day + 1}. No winners selected.")
    else:
        # Выбираем 5 победителей
        winners = random.sample(correct_users, min(5, len(correct_users)))

        # Распределяем подарки
        distribute_gifts(day + 1, winners, context)  # Передаем context в distribute_gifts

        # Отправляем уведомления победителям
        for winner in winners:
            user_id, username = winner
            #if user_id not in notified_winners_global:
                #try:
                    #context.bot.send_message(
                        #chat_id=user_id,
                        #text=f"🎉 Congratulations!\n\nYou are the winner of the day! 🏆✨"
                    #)
                    #logging.info(f"Winner notification sent to user ID: {user_id}")
                    #notified_winners_global.add(user_id)
                #except Exception as e:
                    #logging.error(f"Failed to send winner notification to user ID: {user_id}: {e}")

    wb.save(file_path)
    logging.info(f"Winners for Day {day + 1} have been recorded in the Excel sheet.")

    # Планируем отправку напоминания через 5 секунд
    context.job_queue.run_once(
        notify_users_about_next_day,
        when=5,
        context={'day': day + 1}
    )
    logging.info(f"Reminder for next day scheduled in 5 seconds.")


def distribute_gifts(day, winners, context):
    """
    Распределяет подарки между победителями и обновляет файлы.

    :param day: Текущий день викторины (1-7).
    :param winners: Список победителей в формате [(user_id, username), ...].
    :param context: CallbackContext для отправки сообщений.
    """
    # Загружаем таблицу подарков
    gifts_wb = load_workbook(gifts_file_path)
    gifts_sheet = gifts_wb[f"Day {day}"]

    # Загружаем файл с пользователями
    users_wb = load_workbook(file_path)
    users_sheet = users_wb[f"Day {day}"]

    # Получаем подарки текущего дня
    available_gifts = []
    for row in gifts_sheet.iter_rows(min_row=2, values_only=True):
        gift_name, quantity, remaining = row
        if remaining > 0:
            available_gifts.append((gift_name, remaining))

    if not available_gifts:
        logging.warning(f"No gifts available for Day {day}.")
        return

    for user_id, username in winners:
        if not available_gifts:
            logging.warning(f"All gifts for Day {day} have been distributed.")
            break

        # Выбираем случайный подарок
        selected_gift = random.choice(available_gifts)
        gift_name, remaining = selected_gift

        # Уменьшаем остаток подарка в таблице
        for row in gifts_sheet.iter_rows(min_row=2):
            if row[0].value == gift_name:
                row[2].value -= 1  # Уменьшаем остаток
                break

        # Записываем победителя в файл пользователей
        winner_found = False
        for row in users_sheet.iter_rows(min_row=2):
            if row[0].value == user_id:  # Сравниваем ID пользователя
                row[4].value = "Winner"  # Отмечаем статус
                row[5].value = gift_name  # Название подарка
                winner_found = True
                break

        if not winner_found:
            # Если записи нет, добавляем новую строку
            users_sheet.append([user_id, username, None, None, "Winner", gift_name])

        # Убираем подарок из доступных, если остаток стал 0
        if remaining - 1 == 0:
            available_gifts = [g for g in available_gifts if g[0] != gift_name]

        # Отправляем пользователю сообщение
        context.bot.send_message(
            chat_id=user_id,
            text=f"🎉 Congratulations, {username}! You won: {gift_name} 🎁"
        )

    # Сохраняем обновления
    gifts_wb.save(gifts_file_path)
    users_wb.save(file_path)

    logging.info(f"Gifts for Day {day} distributed and files updated.")
