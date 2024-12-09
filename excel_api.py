import os
from openpyxl import Workbook, load_workbook
import logging
from openpyxl.styles import PatternFill
import csv

# Initialize the Excel file
def initialize_excel(file_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        for i in range(1, 8):
            sheet = wb.create_sheet(title=f"Day {i}")
            headers = ["User ID", "Username", "Response Time", "Correct Answer", "Winner", "Prize"]
            sheet.append(headers)
        wb.remove(wb["Sheet"])
        wb.save(file_path)
        logging.info(f"Excel file initialized with sheets for each quiz day at {file_path}")

# Record user response in Excel
def record_user_response(file_path, user_id, username, day, response_time, result,):
    wb = load_workbook(file_path)
    sheet_name = f"Day {day + 1}"

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    sheet = wb[sheet_name]

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    result_text = "Верно" if result else "Неверно"
    result_fill = green_fill if result else red_fill

    # Проверяем, существует ли уже запись для этого пользователя
    user_found = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == user_id:  # Проверяем по user_id
            row[2].value = response_time  # Обновляем время ответа
            row[3].value = result_text  # Обновляем результат
            row[3].fill = result_fill  # Применяем цвет
            user_found = True
            break

    if not user_found:
        # Если записи нет, добавляем новую
        new_row = [user_id, username, response_time, result_text]
        sheet.append(new_row)

        # Применяем цвет заливки к новой строке
        for cell in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=4):
            if cell[3].value == "Верно":
                cell[3].fill = green_fill
            elif cell[3].value == "Неверно":
                cell[3].fill = red_fill

    wb.save(file_path)
    logging.info(f"Результат для пользователя {username} записан: {result_text}")

    # Function to load authorized usernames from CSV
def load_authorized_usernames(file_path):
    usernames = []
    try:
        with open(file_path, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                if "Telegram Username" in row:
                    usernames.append(row["Telegram Username"])
    except FileNotFoundError:
        logging.warning(f"⚠️ File {file_path} not found. Authorized user list will be empty.")
    except Exception as e:
        logging.error(f"⚠️ Error reading file {file_path}: {e}")
    return usernames